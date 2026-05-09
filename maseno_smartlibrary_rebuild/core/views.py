from datetime import date, timedelta
from io import BytesIO

import openpyxl
import qrcode

from django.db.models import Q, Count, Sum
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from .models import (
    Book,
    Category,
    BorrowRecord,
    Fine,
    SystemSettings,
    Profile,
    AuditLog,
)


def get_settings():
    system_settings = SystemSettings.objects.order_by('-id').first()

    if system_settings is None:
        system_settings = SystemSettings.objects.create(
            school_name="Maseno University",
            library_name="Main Library",
            theme_color="#0d6efd"
        )

    return system_settings


def get_user_profile(user):
    profile, created = Profile.objects.get_or_create(
        user=user,
        defaults={'role': 'student'}
    )
    return profile


def log_action(user, action, description):
    AuditLog.objects.create(
        user=user if user.is_authenticated else None,
        action=action,
        description=description
    )


def admin_or_librarian_required(user):
    return get_user_profile(user).role in ['admin', 'librarian']


def admin_required(user):
    return get_user_profile(user).role == 'admin'


def forbidden_page():
    response = render(None, 'core/access_denied.html', {
        'system_settings': get_settings()
    })
    response.status_code = 403
    return response


def home(request):
    return render(request, 'core/home.html', {
        'system_settings': get_settings()
    })


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None

    if request.method == "POST":
        login_id = request.POST.get('login_id')
        password = request.POST.get('password')

        user = authenticate(request, username=login_id, password=password)

        if user is None:
            try:
                profile = Profile.objects.get(
                    admission_number=login_id,
                    role='student'
                )
                user = authenticate(
                    request,
                    username=profile.user.username,
                    password=password
                )
            except Profile.DoesNotExist:
                user = None

        if user is not None:
            login(request, user)

            log_action(user, 'login', f"{user.username} logged in.")

            profile = get_user_profile(user)

            if profile.must_change_password:
                return redirect('change_password')

            return redirect('dashboard')

        error = "Invalid login details. Students should use admission number. Staff/Admin should use username."

    return render(request, 'core/login.html', {
        'error': error,
        'system_settings': get_settings()
    })


@login_required
def change_password(request):
    profile = get_user_profile(request.user)
    error = None

    if request.method == "POST":
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            error = "Passwords do not match."
        elif len(new_password) < 6:
            error = "Password must be at least 6 characters."
        else:
            request.user.set_password(new_password)
            request.user.save()

            profile.must_change_password = False
            profile.save()

            update_session_auth_hash(request, request.user)

            log_action(
                request.user,
                'password_reset',
                f"{request.user.username} changed their password."
            )

            return redirect('dashboard')

    return render(request, 'core/change_password.html', {
        'system_settings': get_settings(),
        'error': error
    })


@login_required
def dashboard(request):
    profile = get_user_profile(request.user)

    if profile.must_change_password:
        return redirect('change_password')

    if profile.role in ['admin', 'librarian']:
        total_books = Book.objects.count()
        total_users = User.objects.count()
        borrowed = BorrowRecord.objects.filter(returned=False).count()
        overdue = BorrowRecord.objects.filter(
            returned=False,
            due_date__lt=date.today()
        ).count()

        returned_count = BorrowRecord.objects.filter(returned=True).count()
        active_count = BorrowRecord.objects.filter(returned=False).count()

        paid_fines = Fine.objects.filter(paid=True).count()
        unpaid_fines = Fine.objects.filter(paid=False).count()

        total_fines_amount = Fine.objects.aggregate(
            total=Sum('amount')
        )['total'] or 0

        categories = Category.objects.annotate(
            book_count=Count('book')
        ).order_by('name')

        category_labels = [category.name for category in categories]
        category_data = [category.book_count for category in categories]

        return render(request, 'core/dashboard.html', {
            'system_settings': get_settings(),
            'total_books': total_books,
            'total_users': total_users,
            'borrowed': borrowed,
            'overdue': overdue,
            'role': profile.role,
            'returned_count': returned_count,
            'active_count': active_count,
            'paid_fines': paid_fines,
            'unpaid_fines': unpaid_fines,
            'total_fines_amount': total_fines_amount,
            'category_labels': category_labels,
            'category_data': category_data,
        })

    my_records = BorrowRecord.objects.filter(
        user=request.user
    ).order_by('-issue_date')

    my_active = my_records.filter(returned=False)
    my_overdue = my_active.filter(due_date__lt=date.today())
    my_fines = Fine.objects.filter(record__user=request.user)

    return render(request, 'core/student_dashboard.html', {
        'system_settings': get_settings(),
        'profile': profile,
        'role': profile.role,
        'my_records': my_records,
        'my_active': my_active,
        'my_overdue': my_overdue,
        'my_fines': my_fines,
    })


@login_required
def book_list(request):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    query = request.GET.get('q', '').strip()

    books = Book.objects.all().order_by('-created_at')

    if query:
        books = books.filter(
            Q(title__icontains=query) |
            Q(author__icontains=query) |
            Q(isbn__icontains=query) |
            Q(category__name__icontains=query)
        ).distinct()

    return render(request, 'core/book_list.html', {
        'system_settings': get_settings(),
        'books': books,
        'query': query
    })


@login_required
def add_book(request):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    categories = Category.objects.all()

    if request.method == "POST":
        title = request.POST.get('title')
        author = request.POST.get('author')
        category_id = request.POST.get('category')
        total_copies = int(request.POST.get('total_copies'))

        category = Category.objects.get(id=category_id) if category_id else None

        book = Book.objects.create(
            title=title,
            author=author,
            category=category,
            total_copies=total_copies,
            available_copies=total_copies,
            status='available'
        )

        log_action(request.user, 'add_book', f"Added book: {book.title}")

        return redirect('book_list')

    return render(request, 'core/add_book.html', {
        'system_settings': get_settings(),
        'categories': categories
    })


@login_required
def edit_book(request, book_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    book = get_object_or_404(Book, id=book_id)
    categories = Category.objects.all()

    if request.method == "POST":
        old_title = book.title

        book.title = request.POST.get('title')
        book.author = request.POST.get('author')
        category_id = request.POST.get('category')
        book.total_copies = int(request.POST.get('total_copies'))
        book.available_copies = int(request.POST.get('available_copies'))

        book.category = Category.objects.get(id=category_id) if category_id else None
        book.save()

        log_action(request.user, 'edit_book', f"Edited book: {old_title}.")

        return redirect('book_list')

    return render(request, 'core/edit_book.html', {
        'system_settings': get_settings(),
        'book': book,
        'categories': categories
    })


@login_required
def delete_book(request, book_id):
    if not admin_required(request.user):
        return forbidden_page()

    book = get_object_or_404(Book, id=book_id)
    book_title = book.title
    book.delete()

    log_action(request.user, 'delete_book', f"Deleted book: {book_title}")

    return redirect('book_list')


@login_required
def issue_book(request, book_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    book = get_object_or_404(Book, id=book_id)
    users = User.objects.all()

    if book.available_copies <= 0:
        return redirect('book_list')

    if request.method == "POST":
        user_id = request.POST.get('user')
        user = get_object_or_404(User, id=user_id)

        BorrowRecord.objects.create(
            book=book,
            user=user,
            due_date=date.today() + timedelta(days=7)
        )

        book.available_copies -= 1
        book.save()

        log_action(request.user, 'issue_book', f"Issued {book.title} to {user.username}")

        return redirect('borrow_records')

    return render(request, 'core/issue_book.html', {
        'system_settings': get_settings(),
        'book': book,
        'users': users
    })


@login_required
def book_qr_code(request, book_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    book = get_object_or_404(Book, id=book_id)

    issue_url = request.build_absolute_uri(
        reverse('issue_book', args=[book.id])
    )

    qr = qrcode.make(issue_url)

    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="image/png")
    response["Content-Disposition"] = f'inline; filename="book_qr_{book.id}.png"'

    return response


@login_required
def book_qr_print(request, book_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    book = get_object_or_404(Book, id=book_id)

    log_action(
        request.user,
        'issue_book',
        f"Opened QR print page for book: {book.title}"
    )

    return render(request, 'core/book_qr_print.html', {
        'system_settings': get_settings(),
        'book': book
    })


@login_required
def borrow_records(request):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    today = date.today()

    records = BorrowRecord.objects.all().order_by('-issue_date')

    if query:
        records = records.filter(
            Q(book__title__icontains=query) |
            Q(user__username__icontains=query)
        ).distinct()

    if status_filter == "borrowed":
        records = records.filter(returned=False, due_date__gte=today)
    elif status_filter == "overdue":
        records = records.filter(returned=False, due_date__lt=today)
    elif status_filter == "returned":
        records = records.filter(returned=True)

    for record in records:
        end_date = record.return_date if record.returned and record.return_date else today

        if record.due_date < end_date:
            days_overdue = (end_date - record.due_date).days
            amount = days_overdue * 10

            fine, created = Fine.objects.get_or_create(
                record=record,
                defaults={'amount': amount}
            )

            if not fine.paid:
                fine.amount = amount
                fine.save()

    fines = Fine.objects.all()

    return render(request, 'core/borrow_records.html', {
        'system_settings': get_settings(),
        'records': records,
        'today': today,
        'fines': fines,
        'query': query,
        'status_filter': status_filter
    })


@login_required
def return_book(request, record_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    record = get_object_or_404(BorrowRecord, id=record_id)

    if not record.returned:
        record.returned = True
        record.return_date = date.today()
        record.save()

        book = record.book
        book.available_copies += 1
        book.save()

        log_action(request.user, 'return_book', f"Returned {book.title} from {record.user.username}")

    return redirect('borrow_records')


@login_required
def mark_fine_paid(request, fine_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    fine = get_object_or_404(Fine, id=fine_id)
    fine.paid = True
    fine.paid_date = date.today()
    fine.save()

    log_action(
        request.user,
        'fine_paid',
        f"Marked fine paid for {fine.record.user.username}. Amount: KES {fine.amount}"
    )

    return redirect('borrow_records')


@login_required
def export_borrow_records_excel(request):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Borrow Records"

    sheet.append([
        "School",
        "Library",
        "Book",
        "User",
        "Issue Date",
        "Due Date",
        "Return Date",
        "Returned",
        "Fine Amount",
        "Fine Paid",
        "Paid Date"
    ])

    system_settings = get_settings()
    records = BorrowRecord.objects.all().order_by('-issue_date')

    for record in records:
        fine = Fine.objects.filter(record=record).first()

        sheet.append([
            system_settings.school_name,
            system_settings.library_name,
            record.book.title,
            record.user.username,
            str(record.issue_date),
            str(record.due_date),
            str(record.return_date) if record.return_date else "",
            "Yes" if record.returned else "No",
            fine.amount if fine else 0,
            "Yes" if fine and fine.paid else "No",
            str(fine.paid_date) if fine and fine.paid_date else ""
        ])

    log_action(request.user, 'bulk_upload', "Exported borrow records to Excel.")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="borrow_records.xlsx"'

    workbook.save(response)
    return response


@login_required
def borrow_receipt_pdf(request, record_id):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    record = get_object_or_404(BorrowRecord, id=record_id)
    system_settings = get_settings()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="borrow_receipt_{record.id}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 20)
    p.drawCentredString(width / 2, height - 30 * mm, system_settings.school_name)

    p.setFont("Helvetica", 12)
    p.drawCentredString(width / 2, height - 38 * mm, system_settings.library_name)

    p.line(20 * mm, height - 45 * mm, 190 * mm, height - 45 * mm)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(20 * mm, height - 60 * mm, "Library Borrowing Receipt")

    y = height - 78 * mm

    receipt_details = [
        ("Receipt Number", f"BR-{record.id}"),
        ("Borrower", record.user.username),
        ("Book Title", record.book.title),
        ("Author", record.book.author),
        ("Issue Date", str(record.issue_date)),
        ("Due Date", str(record.due_date)),
        ("Returned", "Yes" if record.returned else "No"),
    ]

    for label, value in receipt_details:
        p.setFont("Helvetica-Bold", 11)
        p.drawString(25 * mm, y, f"{label}:")
        p.setFont("Helvetica", 11)
        p.drawString(70 * mm, y, str(value))
        y -= 10 * mm

    p.line(20 * mm, y - 5 * mm, 190 * mm, y - 5 * mm)

    p.setFont("Helvetica-Oblique", 10)
    p.drawString(20 * mm, y - 18 * mm, "Return books before due date to avoid overdue fines.")

    p.setFont("Helvetica", 9)
    p.drawString(20 * mm, 18 * mm, f"Generated by {request.user.username}")

    p.showPage()
    p.save()

    log_action(
        request.user,
        'issue_book',
        f"Generated PDF receipt for {record.user.username} - {record.book.title}"
    )

    return response


@login_required
def add_user(request):
    if not admin_required(request.user):
        return forbidden_page()

    error = None
    success = None

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        admission_number = request.POST.get('admission_number')
        staff_number = request.POST.get('staff_number')
        phone = request.POST.get('phone')

        if User.objects.filter(username=username).exists():
            error = "Username already exists."
        elif role == "student" and not admission_number:
            error = "Admission number is required for students."
        elif role in ["teacher", "librarian"] and not staff_number:
            error = "Staff number is required for staff/librarians."
        else:
            user = User.objects.create_user(username=username, password=password)

            Profile.objects.create(
                user=user,
                role=role,
                admission_number=admission_number if role == "student" else None,
                staff_number=staff_number if role in ["teacher", "librarian"] else None,
                phone=phone
            )

            log_action(request.user, 'add_user', f"Created user {username} with role {role}")

            success = "User created successfully."

    return render(request, 'core/add_user.html', {
        'system_settings': get_settings(),
        'error': error,
        'success': success
    })


@login_required
def bulk_upload_students(request):
    if not admin_required(request.user):
        return forbidden_page()

    error = None
    success = None
    created_count = 0
    skipped_count = 0

    if request.method == "POST":
        excel_file = request.FILES.get("student_file")

        if not excel_file:
            error = "Please upload an Excel file."
        elif not excel_file.name.endswith(".xlsx"):
            error = "Only .xlsx Excel files are allowed."
        else:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name = row[0]
                    admission_number = row[1]
                    phone = row[2] if len(row) > 2 else ""

                    if not name or not admission_number:
                        skipped_count += 1
                        continue

                    username = str(name).strip()
                    admission_number = str(admission_number).strip()
                    phone = str(phone).strip() if phone else ""

                    if Profile.objects.filter(admission_number=admission_number).exists():
                        skipped_count += 1
                        continue

                    if User.objects.filter(username=username).exists():
                        username = f"{username}_{admission_number}"

                    user = User.objects.create_user(
                        username=username,
                        password=admission_number
                    )

                    Profile.objects.create(
                        user=user,
                        role="student",
                        admission_number=admission_number,
                        phone=phone,
                        must_change_password=True
                    )

                    created_count += 1

                log_action(
                    request.user,
                    'bulk_upload',
                    f"Bulk uploaded {created_count} students. Skipped {skipped_count}."
                )

                success = f"{created_count} students uploaded successfully. {skipped_count} skipped."

            except Exception as e:
                error = f"Upload failed: {str(e)}"

    return render(request, "core/bulk_upload_students.html", {
        "system_settings": get_settings(),
        "error": error,
        "success": success
    })


@login_required
def user_list(request):
    if not admin_required(request.user):
        return forbidden_page()

    query = request.GET.get('q', '').strip()

    profiles = Profile.objects.select_related('user').all().order_by(
        'role',
        'user__username'
    )

    if query:
        profiles = profiles.filter(
            Q(user__username__icontains=query) |
            Q(admission_number__icontains=query) |
            Q(staff_number__icontains=query) |
            Q(phone__icontains=query)
        ).distinct()

    return render(request, 'core/user_list.html', {
        'system_settings': get_settings(),
        'profiles': profiles,
        'query': query
    })


@login_required
def edit_user(request, profile_id):
    if not admin_required(request.user):
        return forbidden_page()

    profile = get_object_or_404(Profile, id=profile_id)
    user = profile.user
    error = None

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        admission_number = request.POST.get('admission_number')
        staff_number = request.POST.get('staff_number')
        phone = request.POST.get('phone')
        is_active = request.POST.get('is_active')

        if User.objects.exclude(id=user.id).filter(username=username).exists():
            error = "Username already exists."
        elif role == "student" and not admission_number:
            error = "Admission number is required for students."
        elif role in ["teacher", "librarian"] and not staff_number:
            error = "Staff number is required for staff/librarians."
        else:
            user.username = username
            user.is_active = True if is_active == 'on' else False

            if password:
                user.set_password(password)

            user.save()

            profile.role = role
            profile.admission_number = admission_number if role == "student" else None
            profile.staff_number = staff_number if role in ["teacher", "librarian"] else None
            profile.phone = phone
            profile.save()

            log_action(request.user, 'edit_user', f"Edited user {user.username}")

            return redirect('user_list')

    return render(request, 'core/edit_user.html', {
        'system_settings': get_settings(),
        'profile': profile,
        'error': error
    })


@login_required
def delete_user(request, profile_id):
    if not admin_required(request.user):
        return forbidden_page()

    profile = get_object_or_404(Profile, id=profile_id)

    if profile.user == request.user:
        return redirect('user_list')

    username = profile.user.username
    profile.user.delete()

    log_action(request.user, 'delete_user', f"Deleted user {username}")

    return redirect('user_list')


@login_required
def reset_student_password(request):
    if not admin_or_librarian_required(request.user):
        return forbidden_page()

    error = None
    success = None

    if request.method == "POST":
        admission_number = request.POST.get('admission_number')
        email = request.POST.get('email')

        try:
            profile = Profile.objects.get(
                admission_number=admission_number,
                role='student'
            )

            user = profile.user
            user.email = email
            user.set_password(admission_number)
            user.save()

            profile.must_change_password = True
            profile.save()

            log_action(request.user, 'password_reset', f"Reset password for student {user.username}")

            success = f"Password reset successful. Temporary password is the admission number for {user.username}."

        except Profile.DoesNotExist:
            error = "No student found with that admission number."

    return render(request, 'core/reset_student_password.html', {
        'system_settings': get_settings(),
        'error': error,
        'success': success
    })


@login_required
def audit_logs(request):
    if not admin_required(request.user):
        return forbidden_page()

    logs = AuditLog.objects.select_related('user').all()[:200]

    return render(request, 'core/audit_logs.html', {
        'system_settings': get_settings(),
        'logs': logs
    })


def logout_view(request):
    if request.user.is_authenticated:
        log_action(request.user, 'logout', f"{request.user.username} logged out.")

    logout(request)
    return redirect('login')